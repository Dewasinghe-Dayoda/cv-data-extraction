import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from config import EXPORTS_DIR


EXPORT_COLUMNS = [
    "candidate_name",
    "contact_number",
    "email",
    "gender",
    "age",
    "total_experience_years",
    "last_employer",
    "key_skills",
    "current_job_title",
    "original_filename",
    "created_at"
]

COLUMN_HEADERS = {
    "candidate_name": "Candidate Name",
    "contact_number": "Contact Number",
    "email": "Email Address",
    "gender": "Gender",
    "age": "Age",
    "total_experience_years": "Total Experience (Years)",
    "last_employer": "Last Employer",
    "key_skills": "Key Skills / Expertise",
    "current_job_title": "Current/Last Job Title",
    "original_filename": "Source File",
    "created_at": "Date Extracted"
}


def export_to_excel(candidates: list[dict]) -> str:
    if not candidates:
        raise ValueError("No candidates to export")

    df = pd.DataFrame(candidates)
    for col in EXPORT_COLUMNS:
        if col not in df.columns:
            df[col] = None

    df = df[EXPORT_COLUMNS]
    df.rename(columns=COLUMN_HEADERS, inplace=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"CV_Candidates_{timestamp}.xlsx"
    filepath = EXPORTS_DIR / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "CV Candidates"

    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col_idx, header in enumerate(COLUMN_HEADERS.values(), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    data_font = Font(name="Calibri", size=11)
    data_alignment = Alignment(vertical="center", wrap_text=True)

    for row_idx, candidate in enumerate(candidates, 2):
        for col_idx, col_key in enumerate(EXPORT_COLUMNS, 1):
            value = candidate.get(col_key, "")
            if value is None:
                value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

    column_widths = {
        1: 25, 2: 20, 3: 30, 4: 10, 5: 8,
        6: 18, 7: 25, 8: 40, 9: 25, 10: 20, 11: 18
    }
    for col, width in column_widths.items():
        ws.column_dimensions[chr(64 + col)].width = width

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(str(filepath))
    return str(filepath)
